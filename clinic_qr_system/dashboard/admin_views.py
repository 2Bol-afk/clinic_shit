from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.db import transaction
from patients.models import Patient, StaffProfile, Doctor
from visits.models import Visit, Prescription, LabResult, VaccinationRecord
from .models import AuditLog
import csv
from django.http import HttpResponse
from patients.utils import send_qr_code_email, generate_temp_password


def is_admin(user):
    """Check if user is admin or superuser"""
    return user.is_superuser or user.groups.filter(name='Admin').exists()

def is_doctor(user):
    return user.is_authenticated and user.groups.filter(name='Doctor').exists()

def is_lab(user):
    return user.is_authenticated and user.groups.filter(name='Laboratory').exists()

def is_pharmacy(user):
    return user.is_authenticated and user.groups.filter(name='Pharmacy').exists()

def is_vaccination(user):
    return user.is_authenticated and user.groups.filter(name='Vaccination').exists()

def is_patient(user):
    return user.is_authenticated and hasattr(user, 'patient_profile')

def is_reception(user):
    return user.is_authenticated and user.groups.filter(name='Reception').exists()

@login_required
def reports_redirect(request):
    """Redirect /reports/ to the proper role-specific report."""
    if is_admin(request.user):
        return redirect('admin_system_reports')
    if is_lab(request.user):
        return redirect('lab_report')
    if is_doctor(request.user):
        return redirect('doctor_report')
    if is_pharmacy(request.user):
        # Use the data-backed pharmacy reports view
        return redirect('pharmacy_reports')
    if is_vaccination(request.user):
        return redirect('vaccination_report')
    if is_reception(request.user):
        return redirect('reception_report')
    if is_patient(request.user):
        return redirect('patient_report')
    return redirect('admin_system_reports')

@login_required
@user_passes_test(is_patient)
def patient_report(request):
    start_date = request.GET.get('start_date') or timezone.now().strftime('%Y-%m-%d')
    end_date = request.GET.get('end_date') or (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    # Use timezone-aware datetimes to avoid naive/aware mismatches
    # Use date-based filtering to be robust against naive/aware timezone issues
    # Visits scoped to this patient
    visits = Visit.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date,
        patient__user=request.user
    ).select_related('doctor_user').order_by('-timestamp')
    doctor_visits = visits.filter(service='doctor')
    lab_visits = visits.filter(service='lab')
    prescriptions = Prescription.objects.filter(
        visit__patient__user=request.user,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('visit', 'doctor')
    vaccinations = VaccinationRecord.objects.filter(
        visit__patient__user=request.user,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('visit')
    # Exports
    export = request.GET.get('export')
    if export in ('csv','xlsx','pdf'):
        from django.http import HttpResponse
        filename_base = f"patient_report_{start_date}_to_{end_date}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            
            # CONSULTATIONS SECTION
            w.writerow(['CONSULTATIONS'])
            w.writerow(['Date', 'Doctor', 'Diagnosis/Notes'])
            for v in doctor_visits:
                diagnosis_notes = v.diagnosis or ''
                if v.prescription_notes:
                    diagnosis_notes += f" | Rx: {v.prescription_notes}"
                doctor_name = v.doctor_user.get_full_name() if v.doctor_user else ''
                if v.doctor_user and hasattr(v.doctor_user, 'doctor') and v.doctor_user.doctor.department:
                    doctor_name += f" ({v.doctor_user.doctor.department})"
                w.writerow([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    diagnosis_notes
                ])
            
            # Empty row
            w.writerow([])
            
            # PRESCRIPTIONS SECTION
            w.writerow(['PRESCRIPTIONS'])
            w.writerow(['Date', 'Doctor', 'Medicines', 'Status'])
            for pr in prescriptions:
                try:
                    meds = []
                    for m in pr.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = ' | '.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                doctor_name = pr.doctor.get_full_name() if pr.doctor else ''
                if pr.doctor and hasattr(pr.doctor, 'doctor') and pr.doctor.doctor.department:
                    doctor_name += f" ({pr.doctor.doctor.department})"
                
                w.writerow([
                    pr.created_at.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    medicines_text,
                    pr.get_status_display()
                ])
            
            # Empty row
            w.writerow([])
            
            # LABORATORY TESTS SECTION
            w.writerow(['LABORATORY TESTS'])
            w.writerow(['Date', 'Test', 'Status', 'Results'])
            for v in lab_visits:
                w.writerow([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.lab_test_type or v.lab_tests or 'Lab Test',
                    v.get_status_display(),
                    v.lab_results or ''
                ])
            
            # Empty row
            w.writerow([])
            
            # VACCINATIONS SECTION
            w.writerow(['VACCINATIONS'])
            w.writerow(['Date', 'Vaccine', 'Status'])
            if vaccinations:
                for rec in vaccinations:
                    w.writerow([
                        rec.created_at.strftime('%Y-%m-%d %H:%M'),
                        str(rec.vaccine_type),
                        rec.get_status_display()
                    ])
            else:
                w.writerow(['No vaccination records.'])
            
            return resp
        if export == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('patient_report')
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Patient Report'
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # CONSULTATIONS SECTION
            ws.append(['CONSULTATIONS'])
            ws.append(['Date', 'Doctor', 'Diagnosis/Notes'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).fill = header_fill
            
            for v in doctor_visits:
                diagnosis_notes = v.diagnosis or ''
                if v.prescription_notes:
                    diagnosis_notes += f"\nRx: {v.prescription_notes}"
                doctor_name = v.doctor_user.get_full_name() if v.doctor_user else ''
                if v.doctor_user and hasattr(v.doctor_user, 'doctor') and v.doctor_user.doctor.department:
                    doctor_name += f" ({v.doctor_user.doctor.department})"
                ws.append([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    diagnosis_notes
                ])
            
            # Add empty row
            ws.append([])
            
            # PRESCRIPTIONS SECTION
            ws.append(['PRESCRIPTIONS'])
            ws.append(['Date', 'Doctor', 'Medicines', 'Status'])
            # Apply header styling
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for pr in prescriptions:
                try:
                    meds = []
                    for m in pr.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = '\n'.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                doctor_name = pr.doctor.get_full_name() if pr.doctor else ''
                if pr.doctor and hasattr(pr.doctor, 'doctor') and pr.doctor.doctor.department:
                    doctor_name += f" ({pr.doctor.doctor.department})"
                
                ws.append([
                    pr.created_at.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    medicines_text,
                    pr.get_status_display()
                ])
            
            # Add empty row
            ws.append([])
            
            # LABORATORY TESTS SECTION
            ws.append(['LABORATORY TESTS'])
            ws.append(['Date', 'Test', 'Status', 'Results'])
            # Apply header styling
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for v in lab_visits:
                ws.append([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.lab_test_type or v.lab_tests or 'Lab Test',
                    v.get_status_display(),
                    v.lab_results or ''
                ])
            
            # Add empty row
            ws.append([])
            
            # VACCINATIONS SECTION
            ws.append(['VACCINATIONS'])
            ws.append(['Date', 'Vaccine', 'Status'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            if vaccinations:
                for rec in vaccinations:
                    ws.append([
                        rec.created_at.strftime('%Y-%m-%d %H:%M'),
                        str(rec.vaccine_type),
                        rec.get_status_display()
                    ])
            else:
                ws.append(['No vaccination records.'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO buffer first
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('patient_report')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=A4); width, height = A4
            y = height - 40
            
            # Title and date range
            p.setFont('Helvetica-Bold', 16); p.drawString(40, y, 'Patient Medical Report'); y -= 20
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Report Period: {start_date} to {end_date}'); y -= 30
            
            def draw_section_header(title):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 12); p.drawString(40, y, title); y -= 20
                return y
            
            def draw_table_header(headers):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 10)
                x = 40
                for header, width in headers:
                    p.drawString(x, y, header)
                    x += width
                y -= 15
                # Draw line under headers
                p.line(40, y, x, y)
                y -= 5
                return y
            
            def draw_row(cols, headers):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40
                p.setFont('Helvetica', 9)
                x = 40
                for i, (text, width) in enumerate(zip(cols, [h[1] for h in headers])):
                    # Truncate text to fit column width
                    max_chars = width // 6  # Approximate characters per unit width
                    display_text = str(text)[:max_chars] if text else ''
                    p.drawString(x, y, display_text)
                    x += width
                y -= 12
                return y
            
            # CONSULTATIONS SECTION
            draw_section_header('CONSULTATIONS')
            headers = [('Date', 80), ('Doctor', 120), ('Diagnosis/Notes', 200)]
            draw_table_header(headers)
            
            for v in doctor_visits:
                diagnosis_notes = v.diagnosis or ''
                if v.prescription_notes:
                    diagnosis_notes += f" | Rx: {v.prescription_notes}"
                doctor_name = v.doctor_user.get_full_name() if v.doctor_user else ''
                if v.doctor_user and hasattr(v.doctor_user, 'doctor') and v.doctor_user.doctor.department:
                    doctor_name += f" ({v.doctor_user.doctor.department})"
                draw_row([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    diagnosis_notes
                ], headers)
            
            y -= 20  # Space between sections
            
            # PRESCRIPTIONS SECTION
            draw_section_header('PRESCRIPTIONS')
            headers = [('Date', 80), ('Doctor', 120), ('Medicines', 150), ('Status', 50)]
            draw_table_header(headers)
            
            for pr in prescriptions:
                try:
                    meds = []
                    for m in pr.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = ' | '.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                doctor_name = pr.doctor.get_full_name() if pr.doctor else ''
                if pr.doctor and hasattr(pr.doctor, 'doctor') and pr.doctor.doctor.department:
                    doctor_name += f" ({pr.doctor.doctor.department})"
                
                draw_row([
                    pr.created_at.strftime('%Y-%m-%d %H:%M'),
                    doctor_name,
                    medicines_text,
                    pr.get_status_display()
                ], headers)
            
            y -= 20  # Space between sections
            
            # LABORATORY TESTS SECTION
            draw_section_header('LABORATORY TESTS')
            headers = [('Date', 80), ('Test', 100), ('Status', 60), ('Results', 160)]
            draw_table_header(headers)
            
            for v in lab_visits:
                draw_row([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.lab_test_type or v.lab_tests or 'Lab Test',
                    v.get_status_display(),
                    v.lab_results or ''
                ], headers)
            
            y -= 20  # Space between sections
            
            # VACCINATIONS SECTION
            draw_section_header('VACCINATIONS')
            headers = [('Date', 80), ('Vaccine', 150), ('Status', 70)]
            draw_table_header(headers)
            
            if vaccinations:
                for rec in vaccinations:
                    draw_row([
                        rec.created_at.strftime('%Y-%m-%d %H:%M'),
                        str(rec.vaccine_type),
                        rec.get_status_display()
                    ], headers)
            else:
                draw_row(['No vaccination records.', '', ''], headers)
            
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/patient_report.html', {
        'start_date': start_date, 'end_date': end_date,
        'visits': visits,
        'doctor_visits': doctor_visits,
        'lab_visits': lab_visits,
        'prescriptions': prescriptions,
        'vaccinations': vaccinations,
    })

@login_required
@user_passes_test(is_doctor)
def doctor_report(request):
    start_date = request.GET.get('start_date') or timezone.now().strftime('%Y-%m-%d')
    end_date = request.GET.get('end_date') or (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    # Robust date-only filtering
    visits = Visit.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date,
        service='doctor',
        doctor_user=request.user
    ).order_by('-timestamp')
    prescriptions = Prescription.objects.filter(
        doctor=request.user,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('visit')
    lab_requests = Visit.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date,
        service='lab',
        created_by=request.user
    ).order_by('-timestamp')
    vacc_requests = Visit.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date,
        service='vaccination',
        created_by=request.user
    ).order_by('-timestamp')
    # Exports
    export = request.GET.get('export')
    if export in ('csv','xlsx','pdf'):
        from django.http import HttpResponse
        filename_base = f"doctor_report_{start_date}_to_{end_date}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            
            # CONSULTATIONS SECTION
            w.writerow(['CONSULTATIONS'])
            w.writerow(['Date/Time', 'Patient', 'Status'])
            for v in visits:
                w.writerow([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.patient.full_name if v.patient else '',
                    v.get_status_display()
                ])
            
            # Empty row
            w.writerow([])
            
            # PRESCRIPTIONS SECTION
            w.writerow(['PRESCRIPTIONS'])
            w.writerow(['Date/Time', 'Patient', 'Status'])
            for p in prescriptions:
                w.writerow([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display()
                ])
            
            # Empty row
            w.writerow([])
            
            # LAB REQUESTS SECTION
            w.writerow(['LAB REQUESTS'])
            w.writerow(['Date/Time', 'Patient', 'Status'])
            for l in lab_requests:
                w.writerow([
                    l.timestamp.strftime('%Y-%m-%d %H:%M'),
                    l.patient.full_name if l.patient else '',
                    l.get_status_display()
                ])
            
            # Empty row
            w.writerow([])
            
            # VACCINATION REQUESTS SECTION
            w.writerow(['VACCINATION REQUESTS'])
            w.writerow(['Date/Time', 'Patient', 'Status'])
            for v2 in vacc_requests:
                w.writerow([
                    v2.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v2.patient.full_name if v2.patient else '',
                    v2.get_status_display()
                ])
            
            return resp
        if export == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('doctor_report')
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Doctor Report'
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # CONSULTATIONS SECTION
            ws.append(['CONSULTATIONS'])
            ws.append(['Date/Time', 'Patient', 'Status'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).fill = header_fill
            
            for v in visits:
                ws.append([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.patient.full_name if v.patient else '',
                    v.get_status_display()
                ])
            
            # Add empty row
            ws.append([])
            
            # PRESCRIPTIONS SECTION
            ws.append(['PRESCRIPTIONS'])
            ws.append(['Date/Time', 'Patient', 'Status'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for p in prescriptions:
                ws.append([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display()
                ])
            
            # Add empty row
            ws.append([])
            
            # LAB REQUESTS SECTION
            ws.append(['LAB REQUESTS'])
            ws.append(['Date/Time', 'Patient', 'Status'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for l in lab_requests:
                ws.append([
                    l.timestamp.strftime('%Y-%m-%d %H:%M'),
                    l.patient.full_name if l.patient else '',
                    l.get_status_display()
                ])
            
            # Add empty row
            ws.append([])
            
            # VACCINATION REQUESTS SECTION
            ws.append(['VACCINATION REQUESTS'])
            ws.append(['Date/Time', 'Patient', 'Status'])
            # Apply header styling
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for v2 in vacc_requests:
                ws.append([
                    v2.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v2.patient.full_name if v2.patient else '',
                    v2.get_status_display()
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO buffer first
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('doctor_report')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=A4); width, height = A4
            y = height - 40
            
            # Title and date range
            p.setFont('Helvetica-Bold', 16); p.drawString(40, y, 'Doctor Report'); y -= 20
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Date Range: {start_date} to {end_date}'); y -= 30
            
            def draw_section_header(title):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 12); p.drawString(40, y, title); y -= 20
                return y
            
            def draw_table_header(headers):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 10)
                x = 40
                for header, width in headers:
                    p.drawString(x, y, header)
                    x += width
                y -= 15
                # Draw line under headers
                p.line(40, y, x, y)
                y -= 5
                return y
            
            def draw_row(cols, headers):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40
                p.setFont('Helvetica', 9)
                x = 40
                for i, (text, width) in enumerate(zip(cols, [h[1] for h in headers])):
                    # Truncate text to fit column width
                    max_chars = width // 6  # Approximate characters per unit width
                    display_text = str(text)[:max_chars] if text else ''
                    p.drawString(x, y, display_text)
                    x += width
                y -= 12
                return y
            
            # CONSULTATIONS SECTION
            draw_section_header('CONSULTATIONS')
            headers = [('Date/Time', 80), ('Patient', 200), ('Status', 80)]
            draw_table_header(headers)
            
            for v in visits:
                draw_row([
                    v.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v.patient.full_name if v.patient else '',
                    v.get_status_display()
                ], headers)
            
            y -= 20  # Space between sections
            
            # PRESCRIPTIONS SECTION
            draw_section_header('PRESCRIPTIONS')
            headers = [('Date/Time', 80), ('Patient', 200), ('Status', 80)]
            draw_table_header(headers)
            
            for p in prescriptions:
                draw_row([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display()
                ], headers)
            
            y -= 20  # Space between sections
            
            # LAB REQUESTS SECTION
            draw_section_header('LAB REQUESTS')
            headers = [('Date/Time', 80), ('Patient', 200), ('Status', 80)]
            draw_table_header(headers)
            
            for l in lab_requests:
                draw_row([
                    l.timestamp.strftime('%Y-%m-%d %H:%M'),
                    l.patient.full_name if l.patient else '',
                    l.get_status_display()
                ], headers)
            
            y -= 20  # Space between sections
            
            # VACCINATION REQUESTS SECTION
            draw_section_header('VACCINATION REQUESTS')
            headers = [('Date/Time', 80), ('Patient', 200), ('Status', 80)]
            draw_table_header(headers)
            
            for v2 in vacc_requests:
                draw_row([
                    v2.timestamp.strftime('%Y-%m-%d %H:%M'),
                    v2.patient.full_name if v2.patient else '',
                    v2.get_status_display()
                ], headers)
            
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/doctor_report.html', {
        'start_date': start_date, 'end_date': end_date,
        'visits': visits, 'prescriptions': prescriptions,
        'lab_requests': lab_requests, 'vacc_requests': vacc_requests,
    })

@login_required
@user_passes_test(is_lab)
def lab_report(request):
    start_date = request.GET.get('start_date') or timezone.now().strftime('%Y-%m-%d')
    end_date = request.GET.get('end_date') or (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    # Robust date-only filtering
    lab_visits = Visit.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date,
        service='lab'
    ).select_related('patient').order_by('-timestamp')
    # Verified/In Process from Visit
    verified = lab_visits.filter(status='in_process')
    # Completed from LabResult records to match template expectations (r.visit.*)
    completed = LabResult.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('visit', 'visit__patient').order_by('-created_at')
    # Export handling
    export = request.GET.get('export')
    if export in ('csv','excel','xlsx','pdf'):
        from django.http import HttpResponse
        filename_base = f"laboratory_report_{start_date}_to_{end_date}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            
            # COMPLETED LAB RESULTS SECTION
            w.writerow(['COMPLETED LAB RESULTS'])
            w.writerow(['Patient', 'Patient ID', 'Email', 'Date', 'Test Type', 'Status'])
            for r in completed:
                w.writerow([
                    r.visit.patient.full_name if r.visit and r.visit.patient else '',
                    r.visit.patient.patient_code if r.visit and r.visit.patient else '',
                    r.visit.patient.email if r.visit and r.visit.patient else '',
                    r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                    (r.lab_type or getattr(r.visit, 'lab_test_type', '') or 'Lab Test'),
                    'Done'
                ])
            
            # Empty row
            w.writerow([])
            
            # VERIFIED/IN PROCESS SECTION
            w.writerow(['VERIFIED / IN PROCESS'])
            w.writerow(['Patient', 'Patient ID', 'Email', 'Verified At', 'Test Type', 'Status'])
            for v in verified:
                w.writerow([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    (v.lab_test_type or 'Lab Test'),
                    'In Process'
                ])
            
            return resp
        if export in ('excel','xlsx'):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('lab_report')
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Laboratory Report'
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # COMPLETED LAB RESULTS SECTION
            ws.append(['COMPLETED LAB RESULTS'])
            ws.append(['Patient', 'Patient ID', 'Email', 'Date', 'Test Type', 'Status'])
            # Apply header styling
            for col in range(1, 7):
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).fill = header_fill
            
            for r in completed:
                ws.append([
                    r.visit.patient.full_name if r.visit and r.visit.patient else '',
                    r.visit.patient.patient_code if r.visit and r.visit.patient else '',
                    r.visit.patient.email if r.visit and r.visit.patient else '',
                    r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                    (r.lab_type or getattr(r.visit, 'lab_test_type', '') or 'Lab Test'),
                    'Done'
                ])
            
            # Add empty row
            ws.append([])
            
            # VERIFIED/IN PROCESS SECTION
            ws.append(['VERIFIED / IN PROCESS'])
            ws.append(['Patient', 'Patient ID', 'Email', 'Verified At', 'Test Type', 'Status'])
            # Apply header styling
            for col in range(1, 7):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for v in verified:
                ws.append([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    (v.lab_test_type or 'Lab Test'),
                    'In Process'
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO buffer first
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('lab_report')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=A4); width, height = A4
            y = height - 40
            
            # Title and date range
            p.setFont('Helvetica-Bold', 16); p.drawString(40, y, 'Laboratory Report'); y -= 20
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Date Range: {start_date} to {end_date}'); y -= 30
            
            def draw_section_header(title):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 12); p.drawString(40, y, title); y -= 20
                return y
            
            def draw_table_header(headers):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 10)
                x = 40
                for header, width in headers:
                    p.drawString(x, y, header)
                    x += width
                y -= 15
                # Draw line under headers
                p.line(40, y, x, y)
                y -= 5
                return y
            
            def draw_row(cols, headers):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40
                p.setFont('Helvetica', 9)
                x = 40
                for i, (text, width) in enumerate(zip(cols, [h[1] for h in headers])):
                    # Truncate text to fit column width
                    max_chars = width // 6  # Approximate characters per unit width
                    display_text = str(text)[:max_chars] if text else ''
                    p.drawString(x, y, display_text)
                    x += width
                y -= 12
                return y
            
            # COMPLETED LAB RESULTS SECTION
            draw_section_header('COMPLETED LAB RESULTS')
            headers = [('Patient', 120), ('Patient ID', 80), ('Email', 120), ('Date', 80), ('Test Type', 100), ('Status', 60)]
            draw_table_header(headers)
            
            for r in completed:
                draw_row([
                    r.visit.patient.full_name if r.visit and r.visit.patient else '',
                    r.visit.patient.patient_code if r.visit and r.visit.patient else '',
                    r.visit.patient.email if r.visit and r.visit.patient else '',
                    r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                    (r.lab_type or getattr(r.visit, 'lab_test_type', '') or 'Lab Test'),
                    'Done'
                ], headers)
            
            y -= 20  # Space between sections
            
            # VERIFIED/IN PROCESS SECTION
            draw_section_header('VERIFIED / IN PROCESS')
            headers = [('Patient', 120), ('Patient ID', 80), ('Email', 120), ('Verified At', 80), ('Test Type', 100), ('Status', 60)]
            draw_table_header(headers)
            
            for v in verified:
                draw_row([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    (v.lab_test_type or 'Lab Test'),
                    'In Process'
                ], headers)
            
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/lab_report.html', {
        'start_date': start_date, 'end_date': end_date,
        'verified': verified, 'completed': completed,
    })


@login_required
@user_passes_test(is_reception)
def reception_report(request):
    """Reception report modeled after laboratory layout, but across all visits."""
    start_date = request.GET.get('start_date') or timezone.now().strftime('%Y-%m-%d')
    end_date = request.GET.get('end_date') or (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'), tz)
    end_dt = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'), tz) + timedelta(days=1)

    visits = (Visit.objects
              .filter(timestamp__date__gte=start_date, timestamp__date__lte=end_date, service='reception')
              .select_related('patient')
              .order_by('-timestamp'))

    # Status filtering
    status_filter = (request.GET.get('status') or '').strip()

    in_process = visits.filter(status='in_process')
    completed = visits.filter(status='done')
    # Treat both 'queued' and 'claimed' as part of the queue for visibility
    queued = visits.filter(status__in=['queued', 'claimed'])

    # If a specific status is requested, narrow down for single-section rendering and exports
    if status_filter in ('done', 'in_process', 'queued'):
        if status_filter == 'queued':
            filtered = visits.filter(status__in=['queued', 'claimed'])
        else:
            filtered = visits.filter(status=status_filter)
    else:
        filtered = None

    export = request.GET.get('export')
    if export in ('csv', 'excel', 'xlsx', 'pdf'):
        from django.http import HttpResponse
        filename_base = f"reception_report_{start_date}_to_{end_date}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            
            # COMPLETED VISITS SECTION
            w.writerow(['COMPLETED VISITS'])
            w.writerow(['Patient', 'Patient ID', 'Email', 'Date', 'Service', 'Status'])
            for v in completed:
                w.writerow([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ])
            
            # Empty row
            w.writerow([])
            
            # IN PROCESS SECTION
            w.writerow(['IN PROCESS'])
            w.writerow(['Patient', 'Patient ID', 'Email', 'Started', 'Service', 'Status'])
            for v in in_process:
                w.writerow([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ])
            
            # Empty row
            w.writerow([])
            
            # IN QUEUE SECTION
            w.writerow(['IN QUEUE'])
            w.writerow(['Patient', 'Patient ID', 'Email', 'Date', 'Service', 'Status', 'Queue #'])
            for v in queued:
                w.writerow([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                    v.queue_number or '',
                ])
            
            return resp
        if export in ('excel', 'xlsx'):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('reception_report')
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reception Report'
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # COMPLETED VISITS SECTION
            ws.append(['COMPLETED VISITS'])
            ws.append(['Patient', 'Patient ID', 'Email', 'Date', 'Service', 'Status'])
            # Apply header styling
            for col in range(1, 7):
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).fill = header_fill
            
            for v in completed:
                ws.append([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ])
            
            # Add empty row
            ws.append([])
            
            # IN PROCESS SECTION
            ws.append(['IN PROCESS'])
            ws.append(['Patient', 'Patient ID', 'Email', 'Started', 'Service', 'Status'])
            # Apply header styling
            for col in range(1, 7):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for v in in_process:
                ws.append([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ])
            
            # Add empty row
            ws.append([])
            
            # IN QUEUE SECTION
            ws.append(['IN QUEUE'])
            ws.append(['Patient', 'Patient ID', 'Email', 'Date', 'Service', 'Status', 'Queue #'])
            # Apply header styling
            for col in range(1, 8):
                ws.cell(row=ws.max_row, column=col).font = header_font
                ws.cell(row=ws.max_row, column=col).fill = header_fill
            
            for v in queued:
                ws.append([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                    v.queue_number or '',
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO buffer first
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('reception_report')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=A4); width, height = A4
            y = height - 40
            
            # Title and date range
            p.setFont('Helvetica-Bold', 16); p.drawString(40, y, 'Reception Report'); y -= 20
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Date Range: {start_date} to {end_date}'); y -= 30
            
            def draw_section_header(title):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 12); p.drawString(40, y, title); y -= 20
                return y
            
            def draw_table_header(headers):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 10)
                x = 40
                for header, width in headers:
                    p.drawString(x, y, header)
                    x += width
                y -= 15
                # Draw line under headers
                p.line(40, y, x, y)
                y -= 5
                return y
            
            def draw_row(cols, headers):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40
                p.setFont('Helvetica', 9)
                x = 40
                for i, (text, width) in enumerate(zip(cols, [h[1] for h in headers])):
                    # Truncate text to fit column width
                    max_chars = width // 6  # Approximate characters per unit width
                    display_text = str(text)[:max_chars] if text else ''
                    p.drawString(x, y, display_text)
                    x += width
                y -= 12
                return y
            
            # COMPLETED VISITS SECTION
            draw_section_header('COMPLETED VISITS')
            headers = [('Patient', 120), ('Patient ID', 80), ('Email', 120), ('Date', 80), ('Service', 100), ('Status', 60)]
            draw_table_header(headers)
            
            for v in completed:
                draw_row([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ], headers)
            
            y -= 20  # Space between sections
            
            # IN PROCESS SECTION
            draw_section_header('IN PROCESS')
            headers = [('Patient', 120), ('Patient ID', 80), ('Email', 120), ('Started', 80), ('Service', 100), ('Status', 60)]
            draw_table_header(headers)
            
            for v in in_process:
                draw_row([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                ], headers)
            
            y -= 20  # Space between sections
            
            # IN QUEUE SECTION
            draw_section_header('IN QUEUE')
            headers = [('Patient', 120), ('Patient ID', 80), ('Email', 120), ('Date', 80), ('Service', 100), ('Status', 60), ('Queue #', 50)]
            draw_table_header(headers)
            
            for v in queued:
                draw_row([
                    v.patient.full_name if v.patient else '',
                    v.patient.patient_code if v.patient else '',
                    v.patient.email if v.patient else '',
                    v.timestamp.strftime('%Y-%m-%d %H:%M') if v.timestamp else '',
                    v.get_service_display(),
                    v.get_status_display(),
                    v.queue_number or '',
                ], headers)
            
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/reception_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'in_process': in_process,
        'completed': completed,
        'queued': queued,
        'status_filter': status_filter,
    })

@login_required
@user_passes_test(is_pharmacy)
def pharmacy_report(request):
    # Render pharmacy reports directly with sensible defaults
    start = request.GET.get('start') or request.GET.get('start_date') or (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    end = request.GET.get('end') or request.GET.get('end_date') or timezone.now().strftime('%Y-%m-%d')
    status = request.GET.get('status', '')
    doctor = request.GET.get('doctor', '')

    qs = Prescription.objects.select_related(
        'visit__patient', 'visit__doctor_user', 'doctor', 'dispensed_by'
    ).prefetch_related('medicines').order_by('-created_at')

    if start:
        qs = qs.filter(created_at__date__gte=start)
    if end:
        qs = qs.filter(created_at__date__lte=end)
    if status:
        qs = qs.filter(status=status)
    if doctor:
        qs = qs.filter(doctor__id=doctor)

    stats = {
        'total_prescriptions': qs.count(),
        'pending': qs.filter(status='pending').count(),
        'ready': qs.filter(status='ready').count(),
        'dispensed': qs.filter(status='dispensed').count(),
        'dispensed_today': qs.filter(status='dispensed', dispensed_at__date=timezone.localdate()).count(),
    }

    doctors = User.objects.filter(groups__name='Doctor').order_by('first_name', 'last_name')

    # Export handling
    export = request.GET.get('export')
    if export in ('csv', 'xlsx', 'pdf'):
        from django.http import HttpResponse
        filename_base = f"pharmacy_report_{start}_to_{end}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            
            # PRESCRIPTIONS SECTION
            w.writerow(['PRESCRIPTIONS'])
            w.writerow(['Date', 'Patient', 'Status', 'Medicines', 'Dispensed At'])
            for p in qs:
                try:
                    meds = []
                    for m in p.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = ' | '.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                w.writerow([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display(),
                    medicines_text,
                    p.dispensed_at.strftime('%Y-%m-%d %H:%M') if p.dispensed_at else '',
                ])
            
            return resp
        if export == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('pharmacy_reports')
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Pharmacy Report'
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # PRESCRIPTIONS SECTION
            ws.append(['PRESCRIPTIONS'])
            ws.append(['Date', 'Patient', 'Status', 'Medicines', 'Dispensed At'])
            # Apply header styling
            for col in range(1, 6):
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).fill = header_fill
            
            for p in qs:
                try:
                    meds = []
                    for m in p.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = '\n'.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                ws.append([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display(),
                    medicines_text,
                    p.dispensed_at.strftime('%Y-%m-%d %H:%M') if p.dispensed_at else '',
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO buffer first
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('pharmacy_reports')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=A4); width, height = A4
            y = height - 40
            
            # Title and date range
            p.setFont('Helvetica-Bold', 16); p.drawString(40, y, 'Pharmacy Report'); y -= 20
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Date Range: {start} to {end}'); y -= 30
            
            def draw_section_header(title):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 12); p.drawString(40, y, title); y -= 20
                return y
            
            def draw_table_header(headers):
                nonlocal y
                if y < 60: p.showPage(); y = height - 40
                p.setFont('Helvetica-Bold', 10)
                x = 40
                for header, width in headers:
                    p.drawString(x, y, header)
                    x += width
                y -= 15
                # Draw line under headers
                p.line(40, y, x, y)
                y -= 5
                return y
            
            def draw_row(cols, headers):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40
                p.setFont('Helvetica', 9)
                x = 40
                for i, (text, width) in enumerate(zip(cols, [h[1] for h in headers])):
                    # Truncate text to fit column width
                    max_chars = width // 6  # Approximate characters per unit width
                    display_text = str(text)[:max_chars] if text else ''
                    p.drawString(x, y, display_text)
                    x += width
                y -= 12
                return y
            
            # PRESCRIPTIONS SECTION
            draw_section_header('PRESCRIPTIONS')
            headers = [('Date', 80), ('Patient', 150), ('Status', 80), ('Medicines', 200), ('Dispensed At', 80)]
            draw_table_header(headers)
            
            for p in qs:
                try:
                    meds = []
                    for m in p.medicines.all():
                        med_str = f"• {m.drug_name} {m.dosage}"
                        if m.frequency:
                            med_str += f" {m.frequency}"
                        if m.duration:
                            med_str += f" {m.duration}"
                        meds.append(med_str)
                    medicines_text = ' | '.join(meds) if meds else ''
                except Exception:
                    medicines_text = 'Error loading medicines'
                
                draw_row([
                    p.created_at.strftime('%Y-%m-%d %H:%M'),
                    p.visit.patient.full_name if p.visit and p.visit.patient else '',
                    p.get_status_display(),
                    medicines_text,
                    p.dispensed_at.strftime('%Y-%m-%d %H:%M') if p.dispensed_at else '',
                ], headers)
            
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/pharmacy_reports.html', {
        'prescriptions': qs,
        'stats': stats,
        'doctors': doctors,
        'start': start or '',
        'end': end or '',
        'status': status or '',
        'doctor': doctor or '',
    })

@login_required
@user_passes_test(is_vaccination)
def vaccination_report(request):
    start_date = request.GET.get('start_date') or timezone.now().strftime('%Y-%m-%d')
    end_date = request.GET.get('end_date') or (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    dose_filter = (request.GET.get('dose') or '').strip()
    vaccine_filter = (request.GET.get('vtype') or '').strip()
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    records = VaccinationRecord.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('visit', 'visit__patient').order_by('-created_at')
    # Enrich each record with dose1_date/dose2_date/dose3_date and booster_date from vaccinations app (administered dates only)
    has_booster = False
    try:
        from vaccinations.models import PatientVaccination, VaccineType as VxType
        for r in records:
            r.dose1_date = None
            r.dose2_date = None
            r.dose3_date = None
            r.booster_date = None
            try:
                patient = r.visit.patient
                vname = str(r.vaccine_type)
                vx = VxType.objects.filter(name=vname).first()
                if not vx:
                    continue
                pv = PatientVaccination.objects.filter(patient=patient, vaccine_type=vx).prefetch_related('doses').first()
                if not pv:
                    continue
                for d in pv.doses.all():
                    if not d.administered:
                        continue
                    if d.dose_number == 1:
                        r.dose1_date = d.administered_date
                    elif d.dose_number == 2:
                        r.dose2_date = d.administered_date
                    elif d.dose_number == 3:
                        r.dose3_date = d.administered_date
                    elif d.dose_number and d.dose_number >= 4 and r.booster_date is None:
                        r.booster_date = d.administered_date
                        has_booster = True
            except Exception:
                continue
    except Exception:
        pass
    # If boosters not found via structured doses, fallback to details JSON
    for r in records:
        if r.booster_date:
            continue
        try:
            doses = []
            if isinstance(r.details, dict):
                doses = r.details.get('doses') or []
            if doses:
                for d in doses:
                    label = str(d.get('label','')).lower()
                    if 'booster' in label:
                        date_str = (d.get('date') or '').strip()
                        if date_str:
                            try:
                                r.booster_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                has_booster = True
                                break
                            except Exception:
                                pass
                        else:
                            has_booster = True
                            break
        except Exception:
            continue
    # Apply filters
    def has_in_details(rec, needle: str) -> bool:
        try:
            if isinstance(rec.details, dict):
                doses = rec.details.get('doses') or []
                for d in doses:
                    label = str(d.get('label','')).lower()
                    if needle in label:
                        return True
        except Exception:
            return False
        return False

    filtered_records = []
    for r in records:
        if vaccine_filter and str(r.vaccine_type) != vaccine_filter:
            continue
        if dose_filter == 'dose1' and not (getattr(r, 'dose1_date', None) or has_in_details(r, 'dose 1')):
            continue
        if dose_filter == 'dose2' and not (getattr(r, 'dose2_date', None) or has_in_details(r, 'dose 2')):
            continue
        if dose_filter == 'dose3' and not (getattr(r, 'dose3_date', None) or has_in_details(r, 'dose 3')):
            continue
        if dose_filter == 'booster' and not (getattr(r, 'booster_date', None) or has_in_details(r, 'booster')):
            continue
        filtered_records.append(r)

    # Vaccine type options
    try:
        from visits.models import VaccinationType as VType
        vaccine_types = list(VType.choices)
    except Exception:
        vaccine_types = []
    # Export handling
    export = request.GET.get('export')
    if export in ('csv','xlsx','pdf'):
        from django.http import HttpResponse
        filename_base = f"vaccination_report_{start_date}_to_{end_date}"
        if export == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            w = csv.writer(resp)
            headers = ['Date','Patient','Vaccine','Status','Dose1','Dose2','Dose3'] + (['Booster'] if has_booster else [])
            w.writerow(headers)
            for r in filtered_records:
                row = [
                    r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                    r.visit.patient.full_name if r.visit and r.visit.patient else '',
                    str(r.vaccine_type),
                    r.get_status_display() if hasattr(r,'get_status_display') else r.status,
                    getattr(r,'dose1_date', None) or '',
                    getattr(r,'dose2_date', None) or '',
                    getattr(r,'dose3_date', None) or '',
                ]
                if has_booster:
                    row.append(getattr(r,'booster_date', None) or '')
                w.writerow(row)
            return resp
        if export == 'xlsx':
            try:
                from openpyxl import Workbook
            except Exception:
                messages.error(request, 'XLSX export requires openpyxl')
                return redirect('vaccination_report')
            wb = Workbook(); ws = wb.active; ws.title = 'Vaccination Report'
            headers = ['Date','Patient','Vaccine','Status','Dose1','Dose2','Dose3'] + (['Booster'] if has_booster else [])
            ws.append(headers)
            for r in filtered_records:
                row = [
                    r.created_at,
                    r.visit.patient.full_name if r.visit and r.visit.patient else '',
                    str(r.vaccine_type),
                    r.get_status_display() if hasattr(r,'get_status_display') else r.status,
                    getattr(r,'dose1_date', None) or '',
                    getattr(r,'dose2_date', None) or '',
                    getattr(r,'dose3_date', None) or '',
                ]
                if has_booster:
                    row.append(getattr(r,'booster_date', None) or '')
                ws.append(row)
            # Save to BytesIO buffer first
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return resp
        if export == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.pdfgen import canvas
            except Exception:
                messages.error(request, 'PDF export requires reportlab')
                return redirect('vaccination_report')
            resp = HttpResponse(content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            p = canvas.Canvas(resp, pagesize=landscape(A4)); width, height = landscape(A4)
            y = height - 40
            p.setFont('Helvetica-Bold', 14); p.drawString(40, y, 'Vaccination Report'); y -= 18
            p.setFont('Helvetica', 10); p.drawString(40, y, f'Date Range: {start_date} to {end_date}'); y -= 18
            def draw_row(cols):
                nonlocal y
                if y < 40: p.showPage(); y = height - 40; p.setFont('Helvetica', 10)
                x = 40
                for text, w in cols:
                    p.drawString(x, y, str(text)[:w]); x += 120
                y -= 12
            headers = [('Date',20),('Patient',30),('Vaccine',30),('Status',20),('Dose1',14),('Dose2',14),('Dose3',14)] + (([('Booster',14)]) if has_booster else [])
            p.setFont('Helvetica-Bold', 10); draw_row(headers)
            p.setFont('Helvetica', 9)
            for r in filtered_records:
                row = [
                    (r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '' ,20),
                    ((r.visit.patient.full_name if r.visit and r.visit.patient else ''),30),
                    (str(r.vaccine_type),30),
                    ((r.get_status_display() if hasattr(r,'get_status_display') else r.status),20),
                    (getattr(r,'dose1_date', None) or '',14),
                    (getattr(r,'dose2_date', None) or '',14),
                    (getattr(r,'dose3_date', None) or '',14),
                ]
                if has_booster:
                    row.append((getattr(r,'booster_date', None) or '',14))
                draw_row(row)
            p.showPage(); p.save(); return resp

    return render(request, 'dashboard/vaccination_report.html', {
        'start_date': start_date, 'end_date': end_date,
        'records': filtered_records,
        'has_booster': has_booster,
        'dose_filter': dose_filter,
        'vaccine_filter': vaccine_filter,
        'vaccine_types': vaccine_types,
    })


@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Main admin dashboard with statistics and overview"""
    
    # Get statistics
    stats = {
        'total_patients': Patient.objects.count(),
        'total_staff': StaffProfile.objects.count(),
        'total_doctors': Doctor.objects.count(),
        'total_visits_today': Visit.objects.filter(timestamp__date=timezone.now().date()).count(),
        'total_visits_this_month': Visit.objects.filter(timestamp__month=timezone.now().month).count(),
        'pending_prescriptions': Prescription.objects.filter(status='pending').count(),
        'completed_labs_today': LabResult.objects.filter(created_at__date=timezone.now().date()).count(),
        'vaccinations_today': VaccinationRecord.objects.filter(created_at__date=timezone.now().date()).count(),
    }
    
    # Recent activity
    recent_visits = Visit.objects.select_related('patient').order_by('-timestamp')[:10]
    recent_audit_logs = []
    
    # Department statistics
    # Department statistics (use correct service keys and prescriptions for pharmacy)
    dept_stats = {
        'reception': Visit.objects.filter(service='reception').count(),
        'doctor': Visit.objects.filter(service='doctor').count(),
        'laboratory': Visit.objects.filter(service='lab').count(),
        'vaccination': Visit.objects.filter(service='vaccination').count(),
    }
    try:
        from visits.models import Prescription as _Rx
        # Count all prescriptions in system as pharmacy activity; optionally scope to date if needed
        dept_stats['pharmacy'] = _Rx.objects.count()
    except Exception:
        dept_stats['pharmacy'] = Visit.objects.filter(service='pharmacy').count()
    
    # Date range for reports (default to last 30 days)
    start_date = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    end_date = timezone.now().strftime('%Y-%m-%d')
    
    context = {
        'stats': stats,
        'recent_visits': recent_visits,
        'recent_audit_logs': recent_audit_logs,
        'dept_stats': dept_stats,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'dashboard/admin_dashboard.html', context)










@login_required
@user_passes_test(is_admin)
def patient_management(request):
    """Patient management interface"""
    
    patients = Patient.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        patients = patients.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(patient_code__icontains=search_query) |
            Q(contact__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(patients, 20)
    page_number = request.GET.get('page')
    patients = paginator.get_page(page_number)
    
    context = {
        'patients': patients,
        'search_query': search_query,
    }
    
    return render(request, 'dashboard/patient_management.html', context)


@login_required
@user_passes_test(is_admin)
def resend_qr_code(request, patient_id):
    """Resend QR code to patient email"""
    
    if request.method == 'POST':
        try:
            patient = get_object_or_404(Patient, id=patient_id)
            
            # Generate new QR code and send email
            from patients.utils import generate_qr_code, send_qr_code_email
            
            if generate_qr_code(patient):
                if send_qr_code_email(patient):
                    messages.success(request, 'QR code sent successfully!')
                else:
                    messages.error(request, 'QR code generated but failed to send email!')
            else:
                messages.error(request, 'Failed to generate QR code!')
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='RESEND_QR_CODE',
                details=f'Resent QR code to patient: {patient.full_name}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
        except Exception as e:
            messages.error(request, f'Error sending QR code: {str(e)}')
    
    return redirect('admin_patient_management')


@login_required
@user_passes_test(is_admin)
def audit_logs(request):
    """View audit logs and system activity"""
    
    logs = AuditLog.objects.select_related('user').all()
    
    # Filter by action type
    action_filter = request.GET.get('action')
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    # Filter by user
    user_filter = request.GET.get('user')
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    
    # Order by timestamp (newest first)
    logs = logs.order_by('-timestamp')
    
    # Pagination
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    logs = paginator.get_page(page_number)
    
    # Get available actions for filter
    actions = AuditLog.objects.values_list('action', flat=True).distinct().order_by('action')
    
    context = {
        'logs': logs,
        'actions': actions,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'dashboard/audit_logs.html', context)


@login_required
@user_passes_test(is_admin)
def admin_reports(request):
    """Comprehensive admin reports with department filtering"""
    
    # Date range filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department_filter = request.GET.get('department', 'all')
    
    if not start_date:
        start_date = timezone.now().strftime('%Y-%m-%d')
    if not end_date:
        end_date = (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    
    # Base queryset with date filter
    visits_qs = Visit.objects.filter(timestamp__date__gte=start_date, timestamp__date__lte=end_date)
    
    # Apply department filter
    if department_filter != 'all':
        visits_qs = visits_qs.filter(service=department_filter)
    
    # Department statistics
    # Correct service keys (Visit.Service uses 'lab' not 'laboratory').
    dept_stats = {
        'reception': visits_qs.filter(service='reception').count(),
        'doctor': visits_qs.filter(service='doctor').count(),
        'laboratory': visits_qs.filter(service='lab').count(),
        'vaccination': visits_qs.filter(service='vaccination').count(),
    }
    # Pharmacy commonly uses structured prescriptions instead of Visit('pharmacy').
    try:
        from visits.models import Prescription
        dept_stats['pharmacy'] = Prescription.objects.filter(
            created_at__range=[start_dt, end_dt]
        ).count()
    except Exception:
        dept_stats['pharmacy'] = visits_qs.filter(service='pharmacy').count()
    
    # Staff activity (filtered by department if specified)
    staff_activity = []
    if department_filter == 'all' or department_filter == 'doctor':
        for staff in StaffProfile.objects.all():
            if staff.role == 'doctor':
                doctor_visits = visits_qs.filter(service='doctor', doctor_user=staff.user).count()
                staff_activity.append({
                    'name': staff.user.get_full_name() or staff.user.username,
                    'role': staff.get_role_display(),
                    'visits': doctor_visits
                })
    
    # Department-specific reports
    department_reports = []
    
    if department_filter == 'all' or department_filter == 'doctor':
        # Doctor reports
        doctor_visits = visits_qs.filter(service='doctor')
        department_reports.extend([
            {
                'department': 'Doctor',
                'total_visits': doctor_visits.count(),
                'completed_visits': doctor_visits.filter(status='completed').count(),
                'pending_visits': doctor_visits.filter(status='in_progress').count(),
                'cancelled_visits': doctor_visits.filter(status='cancelled').count(),
            }
        ])
    
    if department_filter == 'all' or department_filter == 'laboratory':
        # Laboratory reports
        lab_visits = visits_qs.filter(service='laboratory')
        department_reports.extend([
            {
                'department': 'Laboratory',
                'total_visits': lab_visits.count(),
                'completed_visits': lab_visits.filter(status='completed').count(),
                'pending_visits': lab_visits.filter(status='in_progress').count(),
                'cancelled_visits': lab_visits.filter(status='cancelled').count(),
            }
        ])
    
    if department_filter == 'all' or department_filter == 'pharmacy':
        # Pharmacy reports
        pharmacy_visits = visits_qs.filter(service='pharmacy')
        department_reports.extend([
            {
                'department': 'Pharmacy',
                'total_visits': pharmacy_visits.count(),
                'completed_visits': pharmacy_visits.filter(status='completed').count(),
                'pending_visits': pharmacy_visits.filter(status='in_progress').count(),
                'cancelled_visits': pharmacy_visits.filter(status='cancelled').count(),
            }
        ])
    
    if department_filter == 'all' or department_filter == 'vaccination':
        # Vaccination reports
        vaccination_visits = visits_qs.filter(service='vaccination')
        department_reports.extend([
            {
                'department': 'Vaccination',
                'total_visits': vaccination_visits.count(),
                'completed_visits': vaccination_visits.filter(status='completed').count(),
                'pending_visits': vaccination_visits.filter(status='in_progress').count(),
                'cancelled_visits': vaccination_visits.filter(status='cancelled').count(),
            }
        ])
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return export_admin_report(visits_qs, dept_stats, staff_activity, export_format, department_filter)
    
    context = {
        'dept_stats': dept_stats,
        'staff_activity': staff_activity,
        'department_reports': department_reports,
        'start_date': start_date,
        'end_date': end_date,
        'department_filter': department_filter,
        'total_visits': visits_qs.count(),
    }
    
    return render(request, 'dashboard/admin_reports.html', context)


def export_admin_report(visits_qs, dept_stats, staff_activity, format_type, department_filter='all'):
    """Export admin report in various formats"""
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = f"admin_report_{department_filter}.csv" if department_filter != 'all' else "admin_report.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Report Type', 'Count'])
        
        # Department statistics
        for dept, count in dept_stats.items():
            writer.writerow([f'{dept.title()} Visits', count])
        
        # Staff activity
        if staff_activity:
            writer.writerow(['', ''])
            writer.writerow(['Staff Activity', ''])
            for staff in staff_activity:
                writer.writerow([f"{staff['name']} ({staff['role']})", staff['visits']])
        
        return response
    
    if format_type in ('excel','xlsx'):
        try:
            from openpyxl import Workbook
        except Exception:
            return HttpResponse('XLSX export requires openpyxl', status=500)
        wb = Workbook(); ws = wb.active; ws.title = 'Admin Report'
        ws.append(['Report Type','Count'])
        for dept, count in dept_stats.items():
            ws.append([f'{dept.title()} Visits', count])
        if staff_activity:
            ws.append([]); ws.append(['Staff Activity',''])
            for s in staff_activity:
                ws.append([f"{s['name']} ({s['role']})", s['visits']])
        # Save to BytesIO buffer first
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"admin_report_{department_filter}.xlsx" if department_filter != 'all' else "admin_report.xlsx"
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except Exception:
            return HttpResponse('PDF export requires reportlab', status=500)
        resp = HttpResponse(content_type='application/pdf')
        filename = f"admin_report_{department_filter}.pdf" if department_filter != 'all' else "admin_report.pdf"
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        p = canvas.Canvas(resp, pagesize=A4); width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14); p.drawString(40, y, 'Admin Report'); y -= 18
        p.setFont('Helvetica', 10); p.drawString(40, y, f'Filter: {department_filter}'); y -= 18
        p.setFont('Helvetica-Bold', 10); p.drawString(40, y, 'Report Type'); p.drawString(240, y, 'Count'); y -= 14
        p.setFont('Helvetica', 10)
        for dept, count in dept_stats.items():
            if y < 40: p.showPage(); y = height - 40; p.setFont('Helvetica', 10)
            p.drawString(40, y, f'{dept.title()} Visits'); p.drawString(240, y, str(count)); y -= 12
        if staff_activity:
            y -= 8; p.setFont('Helvetica-Bold', 10); p.drawString(40, y, 'Staff Activity'); y -= 14; p.setFont('Helvetica', 10)
            for s in staff_activity:
                if y < 40: p.showPage(); y = height - 40; p.setFont('Helvetica', 10)
                p.drawString(40, y, f"{s['name']} ({s['role']})"); p.drawString(300, y, str(s['visits'])); y -= 12
        p.showPage(); p.save(); return resp
    return HttpResponse("Export format not supported", status=400)


@login_required
@user_passes_test(is_admin)
def system_accounts(request):
    """Manage system accounts (Pharmacy, Laboratory, Vaccination)"""
    
    # Get or create system accounts
    system_accounts_data = []
    
    for account_type in ['Pharmacy', 'Laboratory', 'Vaccination', 'Reception']:
        username = f"{account_type.lower()}_account"
        # Try to find existing account
        user = User.objects.filter(username=username).first()
        if not user:
            # Safely get or create to avoid UNIQUE race/inconsistency
            from django.db import IntegrityError
            try:
                user, _ = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'email': f"{account_type.lower()}@clinic.local",
                        'first_name': account_type,
                        'last_name': 'Account',
                        'is_active': True,
                        'password': generate_temp_password(),
                    }
                )
                # If password set via defaults, ensure it is hashed
                if not user.has_usable_password():
                    temp_pw = generate_temp_password()
                    user.set_password(temp_pw)
                    user.is_active = True
                    user.save(update_fields=['password', 'is_active'])
            except IntegrityError:
                # Username exists; fetch it
                user = User.objects.get(username=username)
        # Ensure group membership
        group, _ = Group.objects.get_or_create(name=account_type)
        if not user.groups.filter(id=group.id).exists():
            user.groups.add(group)
        
        system_accounts_data.append({
            'type': account_type,
            'user': user,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_active': user.is_active,
            'last_login': user.last_login,
        })
    
    # Build doctor accounts data with an explicit group membership flag
    doctor_accounts_qs = Doctor.objects.select_related('user').order_by('full_name')
    doctor_accounts = []
    for d in doctor_accounts_qs:
        user = d.user
        has_doctor_group = user.groups.filter(name='Doctor').exists()
        doctor_accounts.append({
            'id': d.id,
            'full_name': d.full_name,
            'specialization': d.specialization,
            'username': user.username,
            'email': user.email,
            'has_doctor_group': has_doctor_group,
        })

    context = {
        'system_accounts': system_accounts_data,
        'doctor_accounts': doctor_accounts,
    }
    
    return render(request, 'dashboard/system_accounts.html', context)


@login_required
@user_passes_test(is_admin)
def edit_system_account(request, account_type):
    """Edit system account details"""
    
    if request.method == 'POST':
        try:
            user = User.objects.get(username=f"{account_type.lower()}_account")
            
            # Update user details
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.is_active = request.POST.get('is_active') == 'on'
            user.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='EDIT_SYSTEM_ACCOUNT',
                details=f'Edited {account_type} account: {user.username}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'{account_type} account updated successfully!')
            
        except User.DoesNotExist:
            messages.error(request, f'{account_type} account not found!')
        except Exception as e:
            messages.error(request, f'Error updating account: {str(e)}')
    
    return redirect('admin_system_accounts')


@login_required
@user_passes_test(is_admin)
def reset_system_account_password(request, account_type):
    """Reset system account password"""
    
    if request.method == 'POST':
        try:
            user = User.objects.get(username=f"{account_type.lower()}_account")
            temp_password = generate_temp_password()
            user.set_password(temp_password)
            user.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='RESET_SYSTEM_ACCOUNT_PASSWORD',
                details=f'Reset password for {account_type} account: {user.username}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'{account_type} account password reset successfully! New password: {temp_password}')
            
        except User.DoesNotExist:
            messages.error(request, f'{account_type} account not found!')
        except Exception as e:
            messages.error(request, f'Error resetting password: {str(e)}')
    
    return redirect('admin_system_accounts')


@login_required
@user_passes_test(is_admin)
def set_system_account_password(request, account_type):
    """Set a specific password for a system account (admin-only)."""
    if request.method == 'POST':
        try:
            user = User.objects.get(username=f"{account_type.lower()}_account")
            pw1 = (request.POST.get('password1') or '').strip()
            pw2 = (request.POST.get('password2') or '').strip()
            if not pw1 or not pw2:
                messages.error(request, 'Password fields cannot be empty.')
                return redirect('admin_system_accounts')
            if pw1 != pw2:
                messages.error(request, 'Passwords do not match.')
                return redirect('admin_system_accounts')
            if len(pw1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return redirect('admin_system_accounts')
            user.set_password(pw1)
            user.save()
            AuditLog.objects.create(
                user=request.user,
                action='SET_SYSTEM_ACCOUNT_PASSWORD',
                details=f'Set password for {account_type} account: {user.username}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            messages.success(request, f'{account_type} account password updated successfully!')
        except User.DoesNotExist:
            messages.error(request, f'{account_type} account not found!')
        except Exception as e:
            messages.error(request, f'Error setting password: {str(e)}')
    return redirect('admin_system_accounts')


@login_required
@user_passes_test(is_admin)
def set_doctor_password(request, doctor_id: int):
    """Set a specific password for a doctor user (admin-only)."""
    if request.method == 'POST':
        try:
            doctor = get_object_or_404(Doctor, pk=doctor_id)
            user = doctor.user
            pw1 = (request.POST.get('password1') or '').strip()
            pw2 = (request.POST.get('password2') or '').strip()
            if not pw1 or not pw2:
                messages.error(request, 'Password fields cannot be empty.')
                return redirect('admin_system_accounts')
            if pw1 != pw2:
                messages.error(request, 'Passwords do not match.')
                return redirect('admin_system_accounts')
            if len(pw1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return redirect('admin_system_accounts')
            user.set_password(pw1)
            user.save()
            messages.success(request, f'Password updated for {doctor.full_name}.')
        except Exception as e:
            messages.error(request, f'Error setting password: {str(e)}')
    return redirect('admin_system_accounts')

@login_required
def system_reports(request):
    """System Reports with role-based filtering"""
    
    # Get filter parameters (may be overridden based on user role)
    role_filter = request.GET.get('role', 'all')
    department_filter = request.GET.get('department', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Set default date range if not provided
    if not start_date:
        start_date = timezone.now().strftime('%Y-%m-%d')
    if not end_date:
        end_date = (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    
    # Base queryset for visits
    visits_qs = Visit.objects.filter(
        timestamp__range=[start_dt, end_dt]
    ).select_related('patient', 'created_by', 'doctor_user', 'doctor_user__doctor_profile')
    
    # Enforce role-based access scoping for non-admin users
    user_groups = set(request.user.groups.values_list('name', flat=True)) if request.user.is_authenticated else set()
    is_superuser = request.user.is_superuser
    if not is_superuser:
        if 'Laboratory' in user_groups:
            role_filter = 'laboratory'
        elif 'Doctor' in user_groups:
            role_filter = 'doctor'
            # Default department to doctor's specialization
            if hasattr(request.user, 'doctor_profile') and request.user.doctor_profile and not department_filter:
                department_filter = request.user.doctor_profile.specialization
        elif 'Pharmacy' in user_groups:
            role_filter = 'pharmacy'
        elif 'Vaccination' in user_groups:
            role_filter = 'vaccination'
        elif 'Patient' in user_groups:
            role_filter = 'patient'
        else:
            # Default to 'all' but no data for unknown roles
            role_filter = 'none'

    # Apply role-based filtering
    reports_data = []
    
    if role_filter == 'all':
        # Show all visits
        visits = visits_qs.order_by('-timestamp')
        for visit in visits:
            # Get department - for doctor visits, use doctor's specialization, otherwise use visit department
            department = 'N/A'
            if visit.service == 'doctor' and visit.doctor_user and hasattr(visit.doctor_user, 'doctor_profile'):
                department = visit.doctor_user.doctor_profile.specialization
            elif visit.department:
                department = visit.department
            
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': visit.get_service_display(),
                'department': department,
                'status': visit.get_status_display(),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': visit.notes or '',
                'queue_number': visit.queue_number or '',
            })
    
    elif role_filter == 'doctor':
        # Filter doctor-related visits
        doctor_visits = visits_qs.filter(service='doctor')
        
        # Apply department filter if specified
        if department_filter and department_filter != '':
            # Filter by doctor's specialization (department) instead of visit department
            doctor_visits = doctor_visits.filter(doctor_user__doctor_profile__specialization=department_filter)
        
        visits = doctor_visits.order_by('-timestamp')
        for visit in visits:
            # Get department from doctor's specialization
            doctor_department = 'N/A'
            if visit.doctor_user and hasattr(visit.doctor_user, 'doctor_profile'):
                doctor_department = visit.doctor_user.doctor_profile.specialization
            
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': 'Consultation',
                'department': doctor_department,
                'status': visit.get_status_display(),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': visit.notes or '',
                'queue_number': visit.queue_number or '',
            })
    
    elif role_filter == 'laboratory':
        # Filter laboratory visits (service value is 'lab')
        lab_visits = visits_qs.filter(service='lab').order_by('-timestamp')
        for visit in lab_visits:
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': 'Lab Test',
                'department': 'Laboratory',
                'status': visit.get_status_display(),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': visit.notes or '',
                'queue_number': visit.queue_number or '',
            })
    
    elif role_filter == 'pharmacy':
        # Prefer structured prescriptions for pharmacy reporting
        try:
            from visits.models import Prescription
            prescriptions = (Prescription.objects
                             .select_related('visit__patient', 'doctor', 'dispensed_by')
                             .filter(created_at__range=[start_dt, end_dt])
                             .order_by('-created_at'))
            for p in prescriptions:
                patient = getattr(p.visit, 'patient', None)
                reports_data.append({
                    'patient_name': patient.full_name if patient else 'N/A',
                    'patient_id': patient.patient_code if patient else 'N/A',
                    'patient_email': patient.email if patient else 'N/A',
                    'date_time': p.created_at,
                    'service_type': 'Prescription',
                    'department': 'Pharmacy',
                    'status': p.get_status_display(),
                    'created_by': (p.doctor.get_full_name() or p.doctor.username) if p.doctor else 'N/A',
                    'doctor': (p.doctor.get_full_name() or p.doctor.username) if p.doctor else 'N/A',
                    'notes': p.pharmacy_notes or '',
                    'queue_number': '',
                })
        except Exception:
            # Fallback to visits if prescriptions are unavailable
            pharmacy_visits = visits_qs.filter(service='pharmacy').order_by('-timestamp')
            for visit in pharmacy_visits:
                reports_data.append({
                    'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                    'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                    'patient_email': visit.patient.email if visit.patient else 'N/A',
                    'date_time': visit.timestamp,
                    'service_type': 'Prescription',
                    'department': 'Pharmacy',
                    'status': visit.get_status_display(),
                    'created_by': visit.created_by.username if visit.created_by else 'N/A',
                    'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                    'notes': visit.notes or '',
                })
    
    elif role_filter == 'vaccination':
        # Filter vaccination visits
        vaccination_visits = visits_qs.filter(service='vaccination').order_by('-timestamp')
        for visit in vaccination_visits:
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': 'Vaccination',
                'department': 'Vaccination',
                'status': visit.get_status_display(),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': visit.notes or '',
                'queue_number': visit.queue_number or '',
            })
    
    elif role_filter == 'reception':
        # Filter reception visits
        reception_visits = visits_qs.filter(service='reception').order_by('-timestamp')
        for visit in reception_visits:
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': 'Reception/Triage',
                'department': visit.department or 'N/A',
                'status': (
                    'Queued' if visit.status == 'queued' else
                    'Claimed' if visit.status == 'claimed' else
                    'In Process' if visit.status == 'in_process' else
                    'Done' if visit.status == 'done' else visit.get_status_display()
                ),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': (visit.notes or ''),
                'queue_number': visit.queue_number or '',
            })
    
    elif role_filter == 'patient':
        # Patient sees only their own records
        patient_visits = visits_qs.filter(patient__user=request.user).order_by('-timestamp')
        for visit in patient_visits:
            # Department for doctor visits via doctor's specialization; otherwise visit.department or service
            department = 'N/A'
            if visit.service == 'doctor' and visit.doctor_user and hasattr(visit.doctor_user, 'doctor_profile'):
                department = visit.doctor_user.doctor_profile.specialization
            elif visit.department:
                department = visit.department
            else:
                department = visit.get_service_display()
            reports_data.append({
                'patient_name': visit.patient.full_name if visit.patient else 'N/A',
                'patient_id': visit.patient.patient_code if visit.patient else 'N/A',
                'patient_email': visit.patient.email if visit.patient else 'N/A',
                'date_time': visit.timestamp,
                'service_type': visit.get_service_display(),
                'department': department,
                'status': visit.get_status_display(),
                'created_by': visit.created_by.username if visit.created_by else 'N/A',
                'doctor': visit.doctor_user.get_full_name() if visit.doctor_user else 'N/A',
                'notes': visit.notes or '',
                'queue_number': visit.queue_number or '',
            })
    
    # Get available departments for doctor role from Patient model choices
    from patients.models import DEPARTMENT_CHOICES
    doctor_departments = DEPARTMENT_CHOICES
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format in ('csv', 'excel', 'pdf', 'xlsx'):
        return export_system_reports_file(reports_data, role_filter, department_filter, start_date, end_date, export_format)
    
    context = {
        'reports_data': reports_data,
        'role_filter': role_filter,
        'department_filter': department_filter,
        'start_date': start_date,
        'end_date': end_date,
        'doctor_departments': doctor_departments,
        'total_records': len(reports_data),
    }
    
    return render(request, 'dashboard/system_reports.html', context)


def export_system_reports_file(reports_data, role_filter, department_filter, start_date, end_date, format_type='csv'):
    """Export system reports as CSV, XLSX, or PDF."""
    if format_type == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO
        except Exception:
            return HttpResponse('XLSX export requires openpyxl', status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'System Reports'
        
        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Headers
        headers = ['Patient Name', 'Patient ID', 'Patient Email', 'Date & Time', 
                  'Service Type', 'Department', 'Queue #', 'Status', 'Created By']
        ws.append(headers)
        
        # Apply header styling
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
        
        # Data rows
        for report in reports_data:
            ws.append([
                report['patient_name'],
                report['patient_id'],
                report['patient_email'],
                report['date_time'].strftime('%Y-%m-%d %H:%M:%S'),
                report['service_type'],
                report['department'],
                report.get('queue_number', ''),
                report['status'],
                report['created_by'],
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO buffer first
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"system_reports_{role_filter}"
        if department_filter:
            filename += f"_{department_filter}"
        filename += f"_{start_date}_to_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    # CSV export
    response = HttpResponse(content_type='text/csv')
    filename = f"system_reports_{role_filter}"
    if department_filter:
        filename += f"_{department_filter}"
    filename += f"_{start_date}_to_{end_date}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Patient Name', 'Patient ID', 'Patient Email', 'Date & Time', 
        'Service Type', 'Department', 'Queue #', 'Status', 'Created By'
    ])
    
    for report in reports_data:
        writer.writerow([
            report['patient_name'],
            report['patient_id'],
            report['patient_email'],
            report['date_time'].strftime('%Y-%m-%d %H:%M:%S'),
            report['service_type'],
            report['department'],
            report.get('queue_number', ''),
            report['status'],
            report['created_by'],
        ])
    
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.colors import Color
        except Exception:
            return HttpResponse('PDF export requires reportlab', status=500)
        resp = HttpResponse(content_type='application/pdf')
        filename = f"system_reports_{role_filter}" + (f"_{department_filter}" if department_filter else '') + f"_{start_date}_to_{end_date}.pdf"
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        p = canvas.Canvas(resp, pagesize=A4); width, height = A4
        y = height - 40
        
        # Title
        p.setFont('Helvetica-Bold', 18); p.drawString(40, y, 'System Reports'); y -= 30
        
        # Summary Section
        p.setFont('Helvetica-Bold', 14); p.drawString(40, y, 'SUMMARY'); y -= 20
        
        # Summary boxes (simulating the card layout)
        box_width = 150
        box_height = 60
        box_y = y - box_height
        
        # Total Records Box
        p.setFillColor(Color(0.13, 0.55, 0.13))  # Dark green background
        p.rect(40, box_y, box_width, box_height, fill=1)
        p.setFillColor(Color(1, 1, 1))  # White text
        p.setFont('Helvetica-Bold', 16); p.drawString(50, box_y + 35, str(len(reports_data)))
        p.setFont('Helvetica', 10); p.drawString(50, box_y + 20, 'Total Records')
        p.setFont('Helvetica', 8); p.drawString(50, box_y + 8, f'{start_date} to {end_date}')
        
        # Role Filter Box
        p.setFillColor(Color(0.0, 0.5, 0.8))  # Blue background
        p.rect(210, box_y, box_width, box_height, fill=1)
        p.setFillColor(Color(1, 1, 1))  # White text
        p.setFont('Helvetica-Bold', 14); p.drawString(220, box_y + 35, role_filter.title())
        p.setFont('Helvetica', 10); p.drawString(220, box_y + 20, 'Role Filter')
        if department_filter:
            p.setFont('Helvetica', 8); p.drawString(220, box_y + 8, department_filter)
        
        # Filter Summary Box
        p.setFillColor(Color(0.9, 0.9, 0.9))  # Light gray background
        p.rect(380, box_y, 200, box_height, fill=1)
        p.setFillColor(Color(0, 0, 0))  # Black text
        p.setFont('Helvetica-Bold', 10); p.drawString(390, box_y + 45, 'Filter Summary')
        p.setFont('Helvetica', 9); 
        summary_text = f'Showing {len(reports_data)} records for {role_filter.title()}'
        if department_filter:
            summary_text += f' in {department_filter}'
        summary_text += f' from {start_date} to {end_date}'
        p.drawString(390, box_y + 25, summary_text[:50])
        if len(summary_text) > 50:
            p.drawString(390, box_y + 15, summary_text[50:100])
        
        y = box_y - 30
        
        # Reports Data Section
        p.setFont('Helvetica-Bold', 14); p.drawString(40, y, 'REPORTS DATA'); y -= 20
        
        # Table headers
        p.setFont('Helvetica-Bold', 10)
        headers = [
            ('Patient Name', 120),
            ('Patient ID', 80),
            ('Patient Email', 120),
            ('Date & Time', 80),
            ('Service Type', 100),
            ('Department', 80),
            ('Queue #', 50),
            ('Status', 80),
            ('Created By', 100)
        ]
        
        x = 40
        for header, width in headers:
            p.drawString(x, y, header)
            x += width
        y -= 15
        
        # Draw line under headers
        p.line(40, y, x, y)
        y -= 10
        
        # Table data
        p.setFont('Helvetica', 8)
        for r in reports_data:
            if y < 60:  # Check if we need a new page
                p.showPage()
                y = height - 40
                p.setFont('Helvetica', 8)
            
            x = 40
            row_data = [
                str(r['patient_name'])[:20],
                str(r['patient_id'])[:12],
                str(r['patient_email'])[:20],
                r['date_time'].strftime('%Y-%m-%d %H:%M'),
                str(r['service_type'])[:15],
                str(r['department'])[:12],
                str(r.get('queue_number', ''))[:8],
                str(r['status'])[:12],
                str(r['created_by'])[:15]
            ]
            
            for i, (text, width) in enumerate(zip(row_data, [h[1] for h in headers])):
                p.drawString(x, y, text)
                x += width
            y -= 12
        
        p.showPage(); p.save(); return resp
    return response
